"""CLI: pull Loads from Alvys, compute per-stop detention, and export a report.

Writes two files:
  - an .xlsx with a "Detention Detail" sheet (autofilter + red highlight on
    flagged rows) and a "By Customer" summary sheet
  - a .json with the same detail rows, in the shape the Detention Tracker
    dashboard (dashboard_template.html) expects
"""

import argparse
import json
import sys
from datetime import date, timedelta

import pandas as pd
from dotenv import load_dotenv
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from alvys_client import AlvysAuthError, AlvysClient
from detention import DETENTION_THRESHOLD_HOURS, filter_rows_by_date, flatten_loads

DETAIL_COLUMNS = [
    "load_number",
    "customer",
    "stop_sequence",
    "stop_type",
    "location",
    "appointment",
    "arrival",
    "departure",
    "hours_from_appointment",
    "dwell_hours",
    "detention_flag",
]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate a per-stop detention report from Alvys loads."
    )
    parser.add_argument("--start-date", default=(date.today() - timedelta(days=30)).isoformat())
    parser.add_argument("--end-date", default=date.today().isoformat())
    parser.add_argument("--customer", default=None, help="Only include loads for this customer.")
    parser.add_argument("--output", default="detention_report.xlsx")
    return parser.parse_args()


def build_rows(start_date, end_date, customer=None):
    client = AlvysClient()
    loads = client.search_all_loads()
    rows = flatten_loads(loads)
    rows = filter_rows_by_date(rows, start_date, end_date)
    if customer:
        needle = customer.strip().lower()
        rows = [r for r in rows if needle in (r["customer"] or "").lower()]
    return rows


def write_excel(rows_df, output_path):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        rows_df.to_excel(writer, sheet_name="Detention Detail", index=False)

        summary = (
            rows_df.groupby("customer")
            .agg(
                stops=("load_number", "count"),
                flagged_stops=("detention_flag", "sum"),
                total_detention_hours=("hours_from_appointment", lambda s: round(s.clip(lower=0).sum(), 2)),
            )
            .sort_values("total_detention_hours", ascending=False)
            .reset_index()
        )
        summary.to_excel(writer, sheet_name="By Customer", index=False)

        detail_sheet = writer.sheets["Detention Detail"]
        detail_sheet.auto_filter.ref = detail_sheet.dimensions
        flag_col_index = DETAIL_COLUMNS.index("detention_flag") + 1
        flag_col_letter = get_column_letter(flag_col_index)
        last_row = len(rows_df) + 1
        red_fill = PatternFill(start_color="FFCFCB", end_color="FFCFCB", fill_type="solid")
        detail_sheet.conditional_formatting.add(
            f"A2:{get_column_letter(len(DETAIL_COLUMNS))}{last_row}",
            CellIsRule(
                operator="equal",
                formula=[f"${flag_col_letter}2=TRUE"],
                fill=red_fill,
            ),
        )
        for col_cells in detail_sheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
            detail_sheet.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

        header_font = Font(bold=True)
        for sheet in (detail_sheet, writer.sheets["By Customer"]):
            for cell in sheet[1]:
                cell.font = header_font


def main():
    load_dotenv()
    args = parse_args()

    try:
        rows = build_rows(args.start_date, args.end_date, customer=args.customer)
    except AlvysAuthError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)

    if not rows:
        print("No stops found for the given date range/filters.", file=sys.stderr)
        sys.exit(0)

    rows_df = pd.DataFrame(rows, columns=DETAIL_COLUMNS)

    write_excel(rows_df, args.output)

    json_path = args.output.rsplit(".", 1)[0] + ".json"
    with open(json_path, "w") as f:
        json.dump(rows, f, indent=2, default=str)

    flagged = sum(1 for r in rows if r["detention_flag"])
    print(f"Detention report written to {args.output} and {json_path}")
    print(f"  Stops: {len(rows)}")
    print(f"  Flagged (> {DETENTION_THRESHOLD_HOURS}h past appointment): {flagged}")


if __name__ == "__main__":
    main()
